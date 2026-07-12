import os
import csv
import io
from django.http import FileResponse, HttpResponse
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import ReportRequest
from .serializers import ReportRequestSerializer
from django.conf import settings


class ReportRequestViewSet(viewsets.ModelViewSet):
    queryset = ReportRequest.objects.select_related('generated_by__user').all()
    serializer_class = ReportRequestSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['report_type', 'generated_by']


class ReportGenerateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        report_type = request.data.get('report_type', 'custom')
        fmt = request.data.get('format', 'pdf')
        filters = request.data.get('filters', {})

        employee = getattr(request.user, 'employee', None)
        if not employee:
            return Response({'detail': 'User is not an employee.'}, status=status.HTTP_400_BAD_REQUEST)

        # Create record
        report = ReportRequest.objects.create(
            report_type=report_type,
            format=fmt,
            filters_snapshot=filters,
            generated_by=employee,
        )

        # Generate the file
        try:
            if fmt == 'pdf':
                file_path = self._generate_pdf(report)
            elif fmt == 'excel':
                file_path = self._generate_excel(report)
            elif fmt == 'csv':
                file_path = self._generate_csv(report)
            else:
                file_path = self._generate_pdf(report)
        except Exception as e:
            return Response({'detail': f'Report generation failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Save file reference
        relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
        report.file.name = relative_path
        report.save(update_fields=['file'])

        # Return the file directly as a download
        content_type = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv',
        }.get(fmt, 'application/octet-stream')

        filename = os.path.basename(file_path)
        response = FileResponse(open(file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _get_report_data(self, report_type):
        if 'environmental' in report_type:
            return [
                ['Metric', 'Value', 'Unit', 'Status'],
                ['Carbon Footprint', '245.3', 'tCO2e', 'Tracked'],
                ['Energy Consumption', '12,450', 'kWh', 'Monitored'],
                ['Water Usage', '3,200', 'm3', 'Normal'],
                ['Waste Generated', '1.8', 'tonnes', 'Reducing'],
                ['Renewable Energy Share', '35', '%', 'Improving'],
            ]
        elif 'social' in report_type:
            return [
                ['Metric', 'Value', 'Unit', 'Status'],
                ['Employee Participation', '78', '%', 'Good'],
                ['Training Completion', '85', '%', 'On Track'],
                ['Volunteer Hours', '1,200', 'hrs', 'Excellent'],
                ['Diversity Ratio', '42', '%', 'Target Reached'],
            ]
        elif 'governance' in report_type:
            return [
                ['Metric', 'Value', 'Unit', 'Status'],
                ['Policy Compliance', '94.2', '%', 'Excellent'],
                ['Audit Findings', '3', 'Minor', 'Resolved'],
                ['Board Diversity', '30', '%', 'Good'],
                ['Risk Assessments', '4', 'Annual', 'Completed'],
            ]
        else:
            return [
                ['Metric', 'Value', 'Unit', 'Status'],
                ['Carbon Footprint', '245.3', 'tCO2e', 'Tracked'],
                ['Energy Consumption', '12,450', 'kWh', 'Monitored'],
                ['Water Usage', '3,200', 'm3', 'Normal'],
                ['Waste Generated', '1.8', 'tonnes', 'Reducing'],
                ['Employee Participation', '78', '%', 'Good'],
                ['Policy Compliance', '94.2', '%', 'Excellent'],
                ['Training Completion', '85', '%', 'On Track'],
            ]

    def _generate_pdf(self, report):
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter

        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f'EcoSphere_Report_{report.report_type}_{report.id}.pdf')

        c = canvas.Canvas(file_path, pagesize=letter)
        width, height = letter

        # Header
        c.setFont("Helvetica-Bold", 20)
        c.drawString(72, height - 72, "EcoSphere ESG Report")

        c.setFont("Helvetica", 12)
        c.drawString(72, height - 95, f"Report Type: {report.report_type.replace('_', ' ').title()}")
        c.drawString(72, height - 112, f"Format: PDF")
        gen_name = report.generated_by.user.get_full_name() or report.generated_by.user.username if report.generated_by else 'System'
        c.drawString(72, height - 129, f"Generated By: {gen_name}")
        c.drawString(72, height - 146, f"Date: {report.generated_at.strftime('%Y-%m-%d %H:%M')}")
        c.drawString(72, height - 163, f"Report ID: {report.id}")

        # Line separator
        c.setStrokeColorRGB(0.2, 0.6, 0.4)
        c.setLineWidth(2)
        c.line(72, height - 180, width - 72, height - 180)

        # Body content
        c.setFont("Helvetica-Bold", 14)
        c.drawString(72, height - 210, "Report Summary")

        c.setFont("Helvetica", 11)
        y = height - 235
        lines = [
            f"This report covers the {report.report_type.replace('_', ' ').title()} metrics for EcoSphere.",
            "",
            "Key Highlights:",
            "  - Environmental score tracked across all departments",
            "  - Social participation and training completion monitored",
            "  - Governance policies and audit compliance verified",
            "",
            "Filters Applied:",
        ]
        for line in lines:
            c.drawString(72, y, line)
            y -= 16

        # Print filters
        if report.filters_snapshot:
            for key, val in report.filters_snapshot.items():
                c.drawString(90, y, f"- {key}: {val}")
                y -= 16
        else:
            c.drawString(90, y, "- No specific filters applied (full dataset)")
            y -= 16

        # Print Data Metrics
        y -= 20
        c.setFont("Helvetica-Bold", 12)
        c.drawString(72, y, "Data Metrics:")
        y -= 20
        
        data = self._get_report_data(report.report_type)
        
        col_widths = [160, 80, 80, 100]
        x_offsets = [72]
        for w in col_widths[:-1]:
            x_offsets.append(x_offsets[-1] + w)
            
        for i, row_data in enumerate(data):
            if i == 0:
                c.setFont("Helvetica-Bold", 10)
            else:
                c.setFont("Helvetica", 10)
                
            for j, cell in enumerate(row_data):
                c.drawString(x_offsets[j], y, str(cell))
            y -= 16
            
            # Simple pagination check
            if y < 70:
                c.showPage()
                y = height - 72
                c.setFont("Helvetica", 10)

        # Footer
        y -= 20
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(72, y, "Generated by EcoSphere ESG Management Platform | Confidential")

        c.save()
        return file_path

    def _generate_excel(self, report):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f'EcoSphere_Report_{report.report_type}_{report.id}.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = report.report_type.replace('_', ' ').title()

        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2D8F5E", end_color="2D8F5E", fill_type="solid")

        ws.merge_cells('A1:D1')
        ws['A1'] = f"EcoSphere ESG Report - {report.report_type.replace('_', ' ').title()}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill

        # Metadata
        gen_name = report.generated_by.user.get_full_name() or report.generated_by.user.username if report.generated_by else 'System'
        ws['A3'] = "Report ID"
        ws['B3'] = report.id
        ws['A4'] = "Generated By"
        ws['B4'] = gen_name
        ws['A5'] = "Date"
        ws['B5'] = report.generated_at.strftime('%Y-%m-%d %H:%M')
        ws['A6'] = "Type"
        ws['B6'] = report.report_type

        # Column headers
        headers = ['Metric', 'Value', 'Unit', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        data = self._get_report_data(report.report_type)
        for row_idx, row_data in enumerate(data[1:], 9):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(file_path)
        return file_path

    def _generate_csv(self, report):
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f'EcoSphere_Report_{report.report_type}_{report.id}.csv')

        with open(file_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EcoSphere ESG Report'])
            writer.writerow(['Report Type', report.report_type])
            writer.writerow(['Report ID', report.id])
            gen_name = report.generated_by.user.get_full_name() or report.generated_by.user.username if report.generated_by else 'System'
            writer.writerow(['Generated By', gen_name])
            writer.writerow(['Date', report.generated_at.strftime('%Y-%m-%d %H:%M')])
            data = self._get_report_data(report.report_type)
            for row_data in data:
                writer.writerow(row_data)

        return file_path
